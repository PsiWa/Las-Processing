from openpyxl.workbook import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


def data_to_excel(df,filename):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    orangeFill = PatternFill(fill_type='solid', start_color='00FFCC00', end_color='00FFCC00')
    blackFill = PatternFill(fill_type='solid', start_color='00000000', end_color='00000000')
    alCenter = Alignment(horizontal='center', vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns)+1)
    #Название шапки
    ws['A1'] = 'Таблица найденных интервалов'
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center',vertical='center')

    ws.cell(row=3, column=1).value = '#'
    ws.cell(row=3, column=1).fill = blackFill
    ws.cell(row=3, column=1).font = Font(bold=True, color='00FFFFFF')
    ws.cell(row=3, column=1).alignment = alCenter
    ws.cell(row=3, column=1).border = thin_border
    for y in range(2,len(df.columns)+2):
        ws.cell(row = 3, column = y).value = df.columns[y-2]
        ws.cell(row=3, column=y).fill = blackFill
        ws.cell(row=3, column=y).font = Font(bold=True,color='00FFFFFF')
        ws.cell(row=3, column=y).alignment = alCenter
        ws.cell(row=3, column=y).border = thin_border
        ws.column_dimensions[get_column_letter(y)].width = len(df.columns[y-2])+10
    for y in range(1, len(df.columns)+2):
        ws.cell(row=4, column=y).value = y
        ws.cell(row=4, column=y).fill = orangeFill
        ws.cell(row=4, column=y).font = Font(bold=True)
        ws.cell(row=4, column=y).alignment = alCenter
        ws.cell(row=4, column=y).border = thin_border
    for x in range(len(df.values)):
        ws.cell(row=x + 5,column=1).value = x+1
        ws.cell(row=x + 5, column=1).border = thin_border
        ws.cell(row=x + 5, column=1).alignment = alCenter
        for y in range(0,len(df.columns)):
            ws.cell(row=x + 5,column=y+2).value = df.iat[x,y]
            ws.cell(row=x + 5, column=y + 2).border = thin_border
            ws.cell(row=x + 5, column=y + 2).alignment = alCenter

    wb.save(f'{filename}.xlsx')

if __name__ == '__main__':
    l1 = [1,1,1]
    l2 = [2,2,2]
    l3 = [3,3,3]
    l4 = [4,4,4]
    l5 = ['s','s','s']
    l6 = [1.1,1.2,1.3]
    df = pd.DataFrame({'Начало интервала': l1,
                       'Конец интервала': l2,
                       'Сдвиг относительно шаблона': l3,
                       'd': l4,
                       'qwer':l5,
                       'qwer':l6})
    data_to_excel(df,'testoutput')


